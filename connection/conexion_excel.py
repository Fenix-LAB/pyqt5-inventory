#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json

class Conexion(object):
    '''
    Clase para manejar la conexión a archivos Excel
    '''

    def __init__(self):
        self.excel_file = 'inventario_perfumeria.xlsx'
        self.workbook = None
        self.current_worksheet = None
        self.cursor = None
        self.db = self  # Para compatibilidad con el código existente
        self.current_data = None
        self.current_headers = None
        self.dataframes = {}  # Cache de DataFrames para cada hoja
        self.lastrowid = 0  # Para compatibilidad con MySQL
        self.setup_excel_file()

    def get_excel_path(self):
        """
        Devuelve la ruta completa del archivo Excel.
        """
        return os.path.abspath(self.excel_file)

    def setup_excel_file(self):
        """
        Configura el archivo Excel. Si no existe, lo crea.
        """
        try:
            if not os.path.exists(self.excel_file):
                print(f"Creando nuevo archivo Excel: {self.excel_file}")
                
                # Crear un nuevo archivo Excel con hojas de ejemplo
                workbook = Workbook()
                
                # Eliminar la hoja por defecto
                default_sheet = workbook.active
                workbook.remove(default_sheet)
                
                # Crear hojas de ejemplo para inventario
                self._create_sample_sheets(workbook)
                
                # Guardar el archivo
                workbook.save(self.excel_file)
                print(f"Archivo Excel creado exitosamente: {self.excel_file}")
            else:
                print(f"Archivo Excel encontrado: {self.excel_file}")
                
        except Exception as e:
            print(f"Error al configurar archivo Excel: {e}")

    def _create_sample_sheets(self, workbook):
        """
        Crea hojas de ejemplo en el archivo Excel
        """
        # Hoja de productos
        productos_sheet = workbook.create_sheet("productos")
        productos_headers = ["id", "nombre", "marca", "categoria", "precio", "stock", "descripcion"]
        productos_sheet.append(productos_headers)
        
        # Hoja de categorias
        categorias_sheet = workbook.create_sheet("categorias")
        categorias_headers = ["id", "nombre", "descripcion"]
        categorias_sheet.append(categorias_headers)
        
        # Datos de ejemplo para categorías
        sample_categories = [
            [1, "Perfumes", "Fragancias para hombre y mujer"],
            [2, "Cremas", "Productos para el cuidado de la piel"],
            [3, "Maquillaje", "Productos de belleza y cosmética"]
        ]
        for row in sample_categories:
            categorias_sheet.append(row)
        
        # Hoja de ventas
        ventas_sheet = workbook.create_sheet("ventas")
        ventas_headers = ["id", "producto_id", "cantidad", "fecha", "total"]
        ventas_sheet.append(ventas_headers)

        # Hoja de clientes
        clientes_sheet = workbook.create_sheet("clientes")
        clientes_headers = ["id", "nombre", "apellido", "email", "telefono", "direccion"]
        clientes_sheet.append(clientes_headers)

        # Hoja de proveedores
        proveedores_sheet = workbook.create_sheet("proveedores")
        proveedores_headers = ["id", "nombre", "contacto", "telefono", "email", "direccion"]
        proveedores_sheet.append(proveedores_headers)

        # Hoja de usuarios
        usuarios_sheet = workbook.create_sheet("usuarios")
        usuarios_headers = ["id", "username", "password", "email", "rol", "activo"]
        usuarios_sheet.append(usuarios_headers)

        # Hoja de marcas
        marcas_sheet = workbook.create_sheet("marcas")
        marcas_headers = ["id", "nombre", "descripcion"]
        marcas_sheet.append(marcas_headers)

        # Hoja de rubros
        rubros_sheet = workbook.create_sheet("rubros")
        rubros_headers = ["id", "nombre", "descripcion"]
        rubros_sheet.append(rubros_headers)

        # Hoja de transacciones
        transacciones_sheet = workbook.create_sheet("transacciones")
        transacciones_headers = ["id", "tipo", "cliente_id", "producto_id", "cantidad", "precio", "fecha", "total"]
        transacciones_sheet.append(transacciones_headers)

        # Hoja de pagos
        pagos_sheet = workbook.create_sheet("pagos")
        pagos_headers = ["id", "transaccion_id", "monto", "metodo_pago", "fecha", "estado"]
        pagos_sheet.append(pagos_headers)

    def conectar(self):
        """
        Conecta con el archivo Excel
        """
        try:
            if os.path.exists(self.excel_file):
                self.workbook = load_workbook(self.excel_file)
                print(f"Conectado al archivo Excel: {self.excel_file}")
                return True
            else:
                print(f"Archivo Excel no encontrado: {self.excel_file}")
                return False
        except Exception as e:
            print(f"Error al conectar con Excel: {e}")
            return False

    def abrir_cursor(self):
        """
        Método de compatibilidad (no hace nada específico en Excel)
        """
        self.cursor = self  # Para mantener compatibilidad con el código existente

    def cerrar_cursor(self):
        """
        Método de compatibilidad (no hace nada específico en Excel)
        """
        self.cursor = None
        self.current_data = None
        self.current_headers = None

    def cerrar_conexion(self):
        """
        Cierra la conexión con Excel y guarda los cambios
        """
        if self.workbook:
            try:
                self.workbook.save(self.excel_file)
                print(f"Cambios guardados en {self.excel_file}")
            except Exception as e:
                print(f"Error al guardar archivo Excel: {e}")
        self.workbook = None
        self.current_worksheet = None
        
    def abrirConexion(self):
        """
        Método para abrir la conexión a Excel (mantiene la compatibilidad)
        """
        self.conectar()
        self.abrir_cursor()
    
    def cerrarConexion(self):
        """
        Método para cerrar la conexión a Excel (mantiene la compatibilidad)
        """
        self.cerrar_cursor()
        self.cerrar_conexion()
    
    def select_worksheet(self, table_name):
        """
        Selecciona una hoja de trabajo específica (equivalente a una tabla en MySQL)
        """
        try:
            if not self.workbook:
                self.conectar()
                
            if table_name in self.workbook.sheetnames:
                self.current_worksheet = self.workbook[table_name]
                # Cargar los datos en un DataFrame para facilitar el trabajo
                self._load_worksheet_data(table_name)
                return True
            else:
                print(f"La hoja '{table_name}' no existe en el archivo Excel.")
                return False
        except Exception as e:
            print(f"Error al seleccionar hoja: {e}")
            return False

    def _load_worksheet_data(self, table_name):
        """
        Carga los datos de la hoja en un DataFrame
        """
        try:
            # Leer la hoja en un DataFrame
            df = pd.read_excel(self.excel_file, sheet_name=table_name)
            self.dataframes[table_name] = df
            
            # Obtener encabezados
            self.current_headers = df.columns.tolist()
        except Exception as e:
            print(f"Error al cargar datos de la hoja {table_name}: {e}")
            # Si falla, crear DataFrame vacío
            self.dataframes[table_name] = pd.DataFrame()
            self.current_headers = []
            
    def execute(self, query, values=None):
        """
        Método para simular la ejecución de consultas SQL en Excel
        query: consulta SQL (se analiza para determinar la acción a realizar)
        values: valores a utilizar en la consulta
        """
        try:
            # Analizar la consulta para determinar qué operación realizar
            query = query.strip().lower()
            
            # Operación SELECT
            if query.startswith("select"):
                return self._handle_select(query, values)
            
            # Operación INSERT
            elif query.startswith("insert"):
                return self._handle_insert(query, values)
                
            # Operación UPDATE
            elif query.startswith("update"):
                return self._handle_update(query, values)
                
            # Operación DELETE
            elif query.startswith("delete"):
                return self._handle_delete(query, values)
                
            else:
                print(f"Operación no soportada: {query}")
                return None
                
        except Exception as e:
            print(f"Error al ejecutar consulta: {e}")
            return None
            
    def _handle_select(self, query, values):
        """
        Maneja consultas SELECT
        """
        try:
            # Extraer nombre de tabla de la consulta
            table_pattern = r"from\s+([a-zA-Z0-9_]+)"
            table_match = re.search(table_pattern, query)
            
            if not table_match:
                print("No se pudo determinar la tabla en la consulta SELECT")
                return None
                
            table_name = table_match.group(1)
            
            # Seleccionar la hoja correspondiente
            if not self.select_worksheet(table_name):
                print(f"No se encontró la hoja '{table_name}'")
                return None
            
            # Obtener DataFrame de la hoja
            df = self.dataframes.get(table_name, pd.DataFrame())
            
            if df.empty:
                self.current_data = []
                return []
            
            # Convertir DataFrame a lista de diccionarios
            all_data = df.to_dict('records')
            
            # Extraer condiciones de la consulta WHERE
            where_pattern = r"where\s+(.*?)(?:$|order\s+by|group\s+by|limit)"
            where_match = re.search(where_pattern, query)
            
            filtered_data = all_data
            
            if where_match:
                conditions = where_match.group(1).strip()
                
                # Convertir valores del parámetro si es necesario
                if values:
                    if isinstance(values, tuple) and len(values) == 1:
                        values = values[0]
                    
                    # Reemplazar comodines %s en las condiciones
                    if isinstance(values, (tuple, list)):
                        for value in values:
                            conditions = conditions.replace("%s", str(value), 1)
                    else:
                        conditions = conditions.replace("%s", str(values))
                
                # Filtrar datos según condiciones
                filtered_data = self._filter_data_by_conditions(all_data, conditions)
            
            self.current_data = filtered_data
            return filtered_data
            
        except Exception as e:
            print(f"Error en consulta SELECT: {e}")
            return None
    
    def _filter_data_by_conditions(self, data, conditions):
        """
        Filtra los datos según las condiciones WHERE (simplificado)
        """
        filtered = []
        
        # Simplificación: solo maneja condiciones básicas como "campo = valor"
        # En un caso real, se requeriría un analizador de condiciones más complejo
        try:
            # Manejo básico para LIKE
            like_pattern = r"([a-zA-Z0-9_\.]+)\s+like\s+['\"]([^'\"]*)['\"]"
            like_match = re.search(like_pattern, conditions)
            
            if like_match:
                field = like_match.group(1).split('.')[-1]  # Eliminar prefijo de tabla si existe
                value_pattern = like_match.group(2).replace('%', '.*')
                
                for row in data:
                    if field in row and re.match(f"^{value_pattern}$", str(row[field]), re.IGNORECASE):
                        filtered.append(row)
                return filtered
                
            # Manejo básico para =, <, >
            condition_pattern = r"([a-zA-Z0-9_\.]+)\s*([=<>])\s*['\"]?([^'\")]*)['\"]?"
            condition_match = re.search(condition_pattern, conditions)
            
            if condition_match:
                field = condition_match.group(1).split('.')[-1]  # Eliminar prefijo de tabla si existe
                operator = condition_match.group(2)
                value = condition_match.group(3)
                
                for row in data:
                    if field in row:
                        if operator == '=' and str(row[field]) == value:
                            filtered.append(row)
                        elif operator == '>' and str(row[field]) > value:
                            filtered.append(row)
                        elif operator == '<' and str(row[field]) < value:
                            filtered.append(row)
                return filtered
                
            # Si no se pudo analizar la condición, devolver todos los datos
            return data
            
        except Exception as e:
            print(f"Error al filtrar datos: {e}")
            return data
    
    def _handle_insert(self, query, values):
        """
        Maneja consultas INSERT
        """
        try:
            # Extraer nombre de tabla y columnas de la consulta
            table_pattern = r"insert\s+into\s+([a-zA-Z0-9_\.]+)\s*\(([^)]+)\)"
            table_match = re.search(table_pattern, query)
            
            if not table_match:
                print("No se pudo determinar la tabla/columnas en la consulta INSERT")
                return None
                
            table_name = table_match.group(1).split('.')[-1]  # Eliminar prefijo de base de datos si existe
            columns = [col.strip() for col in table_match.group(2).split(',')]
            
            # Seleccionar la hoja correspondiente
            if not self.select_worksheet(table_name):
                print(f"No se encontró la hoja '{table_name}'")
                return None
            
            # Verificar que tenemos valores para insertar
            if not values:
                print("No se proporcionaron valores para INSERT")
                return None
            
            # Obtener el DataFrame actual
            df = self.dataframes.get(table_name, pd.DataFrame())
            
            # Preparar la nueva fila
            new_row = {}
            for i, col in enumerate(columns):
                if i < len(values):
                    new_row[col] = values[i]
                else:
                    new_row[col] = ""
            
            # Añadir la nueva fila al DataFrame
            if df.empty:
                # Si el DataFrame está vacío, crear uno nuevo con esta fila
                df = pd.DataFrame([new_row])
            else:
                # Usar pd.concat en lugar de append (que está deprecado)
                new_df = pd.DataFrame([new_row])
                df = pd.concat([df, new_df], ignore_index=True)
            
            # Actualizar el DataFrame en el cache
            self.dataframes[table_name] = df
            
            # Escribir el DataFrame actualizado al archivo Excel
            self._save_dataframe_to_excel(table_name, df)
            
            # Simular la funcionalidad de lastrowid para compatibilidad
            self.lastrowid = len(df)
            
            return True
            
        except Exception as e:
            print(f"Error en consulta INSERT: {e}")
            return None
    
    def _handle_update(self, query, values):
        """
        Maneja consultas UPDATE
        """
        try:
            # Extraer nombre de tabla de la consulta
            table_pattern = r"update\s+([a-zA-Z0-9_\.]+)\s+set"
            table_match = re.search(table_pattern, query)
            
            if not table_match:
                print("No se pudo determinar la tabla en la consulta UPDATE")
                return None
                
            table_name = table_match.group(1).split('.')[-1]  # Eliminar prefijo de base de datos si existe
            
            # Extraer condiciones WHERE
            where_pattern = r"where\s+(.*?)(?:$|order\s+by|group\s+by|limit)"
            where_match = re.search(where_pattern, query)
            
            if not where_match:
                print("No se encontró cláusula WHERE en UPDATE (requerida)")
                return None
                
            conditions = where_match.group(1).strip()
            
            # Extraer campos a actualizar
            set_pattern = r"set\s+(.*?)\s+where"
            set_match = re.search(set_pattern, query)
            
            if not set_match:
                print("No se pudieron determinar los campos a actualizar")
                return None
                
            set_clause = set_match.group(1).strip()
            
            # Seleccionar la hoja correspondiente
            if not self.select_worksheet(table_name):
                print(f"No se encontró la hoja '{table_name}'")
                return None
            
            # Obtener el DataFrame actual
            df = self.dataframes.get(table_name, pd.DataFrame())
            
            if df.empty:
                print(f"No hay datos en la hoja '{table_name}'")
                return False
            
            # Aplicar condición WHERE
            if values:
                if isinstance(values, tuple):
                    # Preparar los valores para el SET y WHERE
                    num_set_values = set_clause.count('%s')
                    set_values = values[:num_set_values]
                    where_values = values[num_set_values:]
                    
                    # Reemplazar los comodines en WHERE
                    conditions_processed = conditions
                    for value in where_values:
                        conditions_processed = conditions_processed.replace("%s", str(value), 1)
                    
                    # Procesar cláusula SET
                    set_pairs = set_clause.split(',')
                    set_fields = []
                    for i, pair in enumerate(set_pairs):
                        if '%s' in pair and i < len(set_values):
                            field = pair.split('=')[0].strip()
                            set_fields.append((field, set_values[i]))
                    
                    # Aplicar filtros para encontrar filas a actualizar
                    all_data = df.to_dict('records')
                    filtered_rows = self._filter_data_by_conditions(all_data, conditions_processed)
                    
                    # Actualizar el DataFrame
                    for row_data in filtered_rows:
                        # Encontrar el índice en el DataFrame
                        mask = pd.Series([True] * len(df))
                        for key, value in row_data.items():
                            if key in df.columns:
                                mask = mask & (df[key] == value)
                        
                        # Actualizar los campos
                        for field, new_value in set_fields:
                            if field in df.columns:
                                df.loc[mask, field] = new_value
                    
                    # Actualizar el DataFrame en el cache
                    self.dataframes[table_name] = df
                    
                    # Escribir el DataFrame actualizado al archivo Excel
                    self._save_dataframe_to_excel(table_name, df)
                    
                    return True
                
        except Exception as e:
            print(f"Error en consulta UPDATE: {e}")
            return None
    
    def _handle_delete(self, query, values):
        """
        Maneja consultas DELETE
        """
        try:
            # Extraer nombre de tabla de la consulta
            table_pattern = r"delete\s+from\s+([a-zA-Z0-9_\.]+)"
            table_match = re.search(table_pattern, query)
            
            if not table_match:
                print("No se pudo determinar la tabla en la consulta DELETE")
                return None
                
            table_name = table_match.group(1).split('.')[-1]  # Eliminar prefijo de base de datos si existe
            
            # Extraer condiciones WHERE
            where_pattern = r"where\s+(.*?)(?:$|order\s+by|group\s+by|limit)"
            where_match = re.search(where_pattern, query)
            
            if not where_match:
                print("No se encontró cláusula WHERE en DELETE (requerida)")
                return None
                
            conditions = where_match.group(1).strip()
            
            # Seleccionar la hoja correspondiente
            if not self.select_worksheet(table_name):
                print(f"No se encontró la hoja '{table_name}'")
                return None
            
            # Obtener el DataFrame actual
            df = self.dataframes.get(table_name, pd.DataFrame())
            
            if df.empty:
                print(f"No hay datos en la hoja '{table_name}'")
                return False
            
            # Aplicar condición WHERE
            if values:
                if isinstance(values, (str, int, float)):
                    # Único valor
                    conditions = conditions.replace("%s", str(values))
                elif isinstance(values, tuple):
                    # Múltiples valores
                    for value in values:
                        conditions = conditions.replace("%s", str(value), 1)
                        
            # Obtener todos los datos y filtrar
            all_data = df.to_dict('records')
            filtered_rows = self._filter_data_by_conditions(all_data, conditions)
            
            # Crear máscara para eliminar filas
            mask_to_keep = pd.Series([True] * len(df))
            for row_data in filtered_rows:
                row_mask = pd.Series([True] * len(df))
                for key, value in row_data.items():
                    if key in df.columns:
                        row_mask = row_mask & (df[key] == value)
                mask_to_keep = mask_to_keep & ~row_mask
            
            # Aplicar la máscara al DataFrame
            df_filtered = df[mask_to_keep]
            
            # Actualizar el DataFrame en el cache
            self.dataframes[table_name] = df_filtered
            
            # Escribir el DataFrame actualizado al archivo Excel
            self._save_dataframe_to_excel(table_name, df_filtered)
                
            return True
                
        except Exception as e:
            print(f"Error en consulta DELETE: {e}")
            return None

    def _save_dataframe_to_excel(self, table_name, df):
        """
        Guarda un DataFrame en una hoja específica del archivo Excel
        """
        try:
            # Cargar el workbook existente
            book = load_workbook(self.excel_file)
            
            # Eliminar la hoja existente si existe
            if table_name in book.sheetnames:
                del book[table_name]
            
            # Crear una nueva hoja
            worksheet = book.create_sheet(table_name)
            
            # Escribir los datos
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
            
            # Guardar el archivo
            book.save(self.excel_file)
            
        except Exception as e:
            print(f"Error al guardar DataFrame en Excel: {e}")
            
    def fetchall(self):
        """
        Devuelve todos los registros de la última consulta
        """
        if self.current_data is not None:
            # Convertir los diccionarios en listas de valores
            result = []
            for row in self.current_data:
                result.append(tuple(row.values()))
            return result
        return []
        
    def fetchone(self):
        """
        Devuelve el primer registro de la última consulta
        """
        if self.current_data and len(self.current_data) > 0:
            return tuple(self.current_data[0].values())
        return None
        
    def commit(self):
        """
        Método de compatibilidad (no es necesario en Excel, se guarda automáticamente)
        """
        pass  # Excel se guarda automáticamente con pandas
