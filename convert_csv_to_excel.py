#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para convertir archivos CSV a hojas del archivo Excel
"""

import pandas as pd
import os
from openpyxl import load_workbook

def convert_csv_to_excel_sheet(csv_file, sheet_name, excel_file='inventario.xlsx'):
    """
    Convierte un archivo CSV a una hoja del archivo Excel
    """
    try:
        # Leer el CSV
        df = pd.read_csv(csv_file)
        print(f"Leyendo {csv_file}...")
        
        # Cargar el workbook existente
        book = load_workbook(excel_file)
        
        # Eliminar la hoja si ya existe
        if sheet_name in book.sheetnames:
            del book[sheet_name]
            print(f"Hoja '{sheet_name}' eliminada")
        
        # Crear una nueva hoja
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"✅ Datos de {csv_file} importados a la hoja '{sheet_name}'")
        
    except Exception as e:
        print(f"❌ Error al convertir {csv_file}: {e}")

if __name__ == "__main__":
    # Ejemplo de uso
    # convert_csv_to_excel_sheet('productos.csv', 'productos')
    # convert_csv_to_excel_sheet('clientes.csv', 'clientes')
    
    print("Script de conversión CSV a Excel")
    print("Uso: convert_csv_to_excel_sheet('archivo.csv', 'nombre_hoja')")
