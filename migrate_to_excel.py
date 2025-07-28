#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script de migración de Google Sheets a Excel
Convierte los datos existentes de Google Sheets al nuevo formato Excel
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook

def migrate_to_excel():
    """
    Migra los datos de Google Sheets a Excel
    """
    print("=== MIGRACIÓN DE GOOGLE SHEETS A EXCEL ===")
    print()
    
    # Verificar si existe el archivo de credenciales de Google Sheets
    credentials_file = 'credentials.json'
    if os.path.exists(credentials_file):
        print(f"Se encontró el archivo de credenciales: {credentials_file}")
        print("Si tienes datos en Google Sheets, puedes migrarlos manualmente.")
        print("Para esto, descarga cada hoja como CSV desde Google Sheets")
        print("y luego usa el script convert_csv_to_excel.py")
    else:
        print("No se encontró archivo de credenciales de Google Sheets.")
        print("Creando archivo Excel desde cero...")
    
    # Crear el nuevo archivo Excel
    excel_file = 'inventario_perfumeria.xlsx'
    
    if os.path.exists(excel_file):
        print(f"El archivo {excel_file} ya existe.")
        response = input("¿Deseas sobrescribirlo? (s/n): ")
        if response.lower() != 's':
            print("Migración cancelada.")
            return
    
    # Crear nuevo workbook
    workbook = Workbook()
    
    # Eliminar la hoja por defecto
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Definir las hojas y sus estructuras
    sheets_structure = {
        "productos": ["id", "nombre", "marca", "categoria", "precio", "stock", "descripcion"],
        "categorias": ["id", "nombre", "descripcion"],
        "clientes": ["id", "nombre", "apellido", "email", "telefono", "direccion"],
        "proveedores": ["id", "nombre", "contacto", "telefono", "email", "direccion"],
        "usuarios": ["id", "username", "password", "email", "rol", "activo"],
        "marcas": ["id", "nombre", "descripcion"],
        "rubros": ["id", "nombre", "descripcion"],
        "ventas": ["id", "producto_id", "cantidad", "fecha", "total"],
        "transacciones": ["id", "tipo", "cliente_id", "producto_id", "cantidad", "precio", "fecha", "total"],
        "pagos": ["id", "transaccion_id", "monto", "metodo_pago", "fecha", "estado"]
    }
    
    # Crear cada hoja con sus encabezados
    for sheet_name, headers in sheets_structure.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        print(f"Creada hoja: {sheet_name}")
    
    # Agregar algunos datos de ejemplo en categorías
    categorias_sheet = workbook["categorias"]
    sample_categories = [
        [1, "Perfumes", "Fragancias para hombre y mujer"],
        [2, "Cremas", "Productos para el cuidado de la piel"],
        [3, "Maquillaje", "Productos de belleza y cosmética"]
    ]
    for row in sample_categories:
        categorias_sheet.append(row)
    
    # Guardar el archivo
    workbook.save(excel_file)
    print(f"\n✅ Archivo Excel creado exitosamente: {excel_file}")
    print()
    print("SIGUIENTE PASOS:")
    print("1. Instala las dependencias: pip install -r requirements.txt")
    print("2. Ejecuta tu aplicación normalmente")
    print("3. Los datos se guardarán automáticamente en el archivo Excel")
    print()
    print("Si tienes datos en Google Sheets:")
    print("1. Descarga cada hoja como CSV desde Google Sheets")
    print("2. Usa pandas para convertir los CSV a hojas de Excel")
    print("3. O importa manualmente los datos usando la aplicación")

def create_csv_converter():
    """
    Crea un script auxiliar para convertir CSV a Excel
    """
    converter_script = '''#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para convertir archivos CSV a hojas del archivo Excel
"""

import pandas as pd
import os
from openpyxl import load_workbook

def convert_csv_to_excel_sheet(csv_file, sheet_name, excel_file='inventario_perfumeria.xlsx'):
    """
    Convierte un archivo CSV a una hoja del archivo Excel
    """
    try:
        # Leer el CSV
        df = pd.read_csv(csv_file)
        print(f"Leyendo {csv_file}...")
        
        # Cargar el workbook existente
        book = load_workbook(excel_file)
        
        # Eliminar la hoja si ya existe
        if sheet_name in book.sheetnames:
            del book[sheet_name]
            print(f"Hoja '{sheet_name}' eliminada")
        
        # Crear una nueva hoja
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"✅ Datos de {csv_file} importados a la hoja '{sheet_name}'")
        
    except Exception as e:
        print(f"❌ Error al convertir {csv_file}: {e}")

if __name__ == "__main__":
    # Ejemplo de uso
    # convert_csv_to_excel_sheet('productos.csv', 'productos')
    # convert_csv_to_excel_sheet('clientes.csv', 'clientes')
    
    print("Script de conversión CSV a Excel")
    print("Uso: convert_csv_to_excel_sheet('archivo.csv', 'nombre_hoja')")
'''
    
    with open('convert_csv_to_excel.py', 'w', encoding='utf-8') as f:
        f.write(converter_script)
    
    print("✅ Script auxiliar creado: convert_csv_to_excel.py")

if __name__ == "__main__":
    migrate_to_excel()
    create_csv_converter()
